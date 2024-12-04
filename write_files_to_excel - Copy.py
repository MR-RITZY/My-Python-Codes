import openpyxl, argparse
from openpyxl.utils import get_column_letter
parser = argparse.ArgumentParser()
parser.add_argument('spreadsheet')
parser.add_argument('filenames', nargs = '+')
arg = parser.parse_args()
wbook = openpyxl.Workbook()
sheet = wbook.active
for k, i in enumerate(arg.filenames):
    fil = open(i)
    file_text = fil.readlines()
    for j, l in enumerate(file_text):
        sheet[get_column_letter(k + 1) + str(j + 1)] = l
wbook.save(arg.spreadsheet)

