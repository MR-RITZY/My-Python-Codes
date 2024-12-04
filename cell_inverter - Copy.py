import openpyxl, argparse
from openpyxl.utils import get_column_letter, column_index_from_string
parser = argparse.ArgumentParser()
parser.add_argument('filename')
parser.add_argument('--sheet')
arg = parser.parse_args()
workbook = openpyxl.load_workbook(arg.filename)
if arg.sheet is not None:
    worksheet = workbook[arg.sheet]
else:
    worksheet = workbook.active
trans_sheet = workbook.create_sheet(title = f'trans_{worksheet.title}')
for i in range(1, worksheet.max_column + 1):
    for j in range(1, worksheet.max_row + 1):
        trans_sheet[get_column_letter(i) + str(j)] = worksheet[get_column_letter(j) + str(i)].value
workbook.save(arg.filename)
