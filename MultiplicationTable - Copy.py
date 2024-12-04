import openpyxl, argparse
from openpyxl.utils import get_column_letter
parser = argparse.ArgumentParser()
parser.add_argument('number', type = int)
arg = parser.parse_args()
workbook = openpyxl.Workbook()
sheet = workbook.active
for i in range(1, arg.number+1):
    sheet['A' + str(i+1)] = i
    sheet[get_column_letter(i+1) + '1'] = i
for i in range(1, arg.number+1):
    for j in range(1, arg.number+1):
        sheet[get_column_letter(i + 1) + str(j + 1)] = f'=PRODUCT(A{i+1}, {get_column_letter(j+1)}1)'
workbook.save(f'Multiplication Table {arg.number}.xlsx')

